"""
Load field bus data for UXsim calibration.

Source: Bus_Data_Collection_June 2026.xlsx (June 2026 — permissioned San Agustin study).

Per-bus simulation uses San Agustin stationary rows only; other companies are
background corridor traffic (noise), not individually modeled.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Literal

import openpyxl

import corridor_config as corridor_config
from corridor_config import (
    DEFAULT_DWELL_S,
    SESSION_BUS_FIELD,
    SESSION_BUS_FIELD_SIM,
    SESSION_BUS_PER_HOUR_BY_DIRECTION,
    SESSION_ONBOARD_JAN18,
    SESSION_ONBOARD_TRAVEL_S,
    SessionBusField,
)

_PROJECT = Path(__file__).resolve().parent
JUNE_2026_DATA_PATH = _PROJECT / "Bus_Data_Collection_June 2026.xlsx"


def require_june2026_workbook() -> Path:
    if not JUNE_2026_DATA_PATH.is_file():
        raise FileNotFoundError(
            f"Missing field workbook: {JUNE_2026_DATA_PATH}. "
            "Place Bus_Data_Collection_June 2026.xlsx in the project folder."
        )
    return JUNE_2026_DATA_PATH


FIELD_DATA_PATH = require_june2026_workbook()
DATA_PATH = FIELD_DATA_PATH

# Permissioned study operator — only this company is simulated per-bus.
SIMULATION_BUS_COMPANY_KEY = "san agustin"


def study_operator_label() -> str:
    return SIMULATION_BUS_COMPANY_KEY.title()


def is_study_bus_company(name: str) -> bool:
    return SIMULATION_BUS_COMPANY_KEY in str(name or "").strip().lower()


def is_erjohn_bus_company(name: str) -> bool:
    """Deprecated alias — use is_study_bus_company()."""
    return is_study_bus_company(name)


def is_simulation_bus_company(name: str, *, study_operator_only: bool = True) -> bool:
    if not study_operator_only:
        return bool(str(name or "").strip())
    return is_study_bus_company(name)


def field_workbook_is_june2026() -> bool:
    return True


SESSION_SHEETS = ("Morning Session", "Lunch", "Afternoon Session")

DataOrigin = Literal["robinsons", "waltermart"]

SIM_CLEARANCE_S = 10800.0
MIN_DEMAND_SPAN_S = 1800.0
MAX_DEMAND_WINDOW_CAP_S = 5 * 3600.0

# Stationary row time-of-day -> session label (matches corridor_config.SESSIONS).
_SESSION_HOURS: tuple[tuple[str, int, int], ...] = (
    ("Morning Session", 5, 10),
    ("Lunch", 11, 14),
    ("Afternoon Session", 15, 20),
)


def _time_to_seconds(val):
    if val is None:
        return None
    if isinstance(val, time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, datetime):
        return _time_to_seconds(val.time())
    s = str(val).strip()
    m = re.match(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", s)
    if m:
        h, mn, sec = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        return h * 3600 + mn * 60 + sec
    return None


def _session_from_clock(t: time | None) -> str | None:
    if t is None:
        return None
    h = t.hour
    for name, lo, hi in _SESSION_HOURS:
        if lo <= h < hi:
            return name
    return None


def _normalize_location(loc: str) -> str:
    t = (loc or "").strip().lower()
    if "robinson" in t:
        return "robinsons"
    if "walter" in t:
        return "waltermart"
    return "other"


def parse_dwell_seconds(raw):
    """Parse dwell cells like '30 s', '1 m 35 s', '2m 10s'."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().lower()
    total = 0.0
    ok = False
    for m in re.finditer(r"(\d+)\s*m", s):
        total += int(m.group(1)) * 60
        ok = True
    for m in re.finditer(r"(\d+)\s*s", s):
        total += int(m.group(1))
        ok = True
    return total if ok else None


def parse_travel_time_seconds(raw) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    m = re.search(r"(\d+)\s*min", s)
    if m:
        return float(m.group(1)) * 60.0
    return parse_dwell_seconds(raw)


def parse_signal_delay_seconds(raw) -> float | None:
    """Use waiting time before 'delayed' (total wait at signals on the trip)."""
    if raw is None:
        return None
    s = str(raw).split("delayed")[0]
    return parse_dwell_seconds(s)


def parse_speed_range_mps(raw) -> float | None:
    """Midpoint of '7 km/h - 29 km/h' style on-board cells."""
    if raw is None:
        return None
    s = str(raw).lower().replace("km/h", "").replace("kmh", "")
    nums = [float(x) for x in re.findall(r"[\d.]+", s)]
    if not nums:
        return None
    return sum(nums) / len(nums) / 3.6


_ONBOARD_COL_TO_SESSION = {
    "6:00 - 7:00 AM": "Morning Session",
    "12:00 - 1:00 PM": "Lunch",
    "4:00 - 5:00 PM": "Afternoon Session",
}

_ONBOARD_ROW_KEYS = {
    "dwell": ("dwell time (bus stop)", "dwell time"),
    "signal": ("traffic signal delay", "signal delay", "time waiting"),
    "speed": ("average speed",),
    "travel": ("travel time to destination",),
    "crowding": ("crowding level",),
    "traffic": ("traffic condition",),
}


def _is_onboard_sheet(name: str) -> bool:
    return str(name or "").strip().lower().startswith("onboard")


def _onboard_row_map(rows: list) -> dict[str, int]:
    for row in rows[:6]:
        if row and str(row[0] or "").strip().lower().startswith("parameters"):
            return {str(row[i]).strip(): i for i in range(1, len(row)) if row[i]}
    return {}


def _find_onboard_row(rows: list, *needles: str) -> tuple | None:
    for row in rows:
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip().lower()
        if any(n in label for n in needles):
            return row
    return None


def _onboard_bus_name_from_notes(rows: list, col_idx: int) -> str:
    row_n = _find_onboard_row(rows, "notes")
    if not row_n or col_idx >= len(row_n):
        return ""
    m = re.search(r"name of bus:\s*(.+)", str(row_n[col_idx] or ""), re.I)
    return m.group(1).strip() if m else ""


def load_onboard_session_records(
    *,
    study_operator_only: bool = False,
) -> list[dict[str, object]]:
    """One dict per Onboard sheet column (session slot), with crowding/traffic labels."""
    if not FIELD_DATA_PATH.is_file():
        return []
    wb = openpyxl.load_workbook(FIELD_DATA_PATH, data_only=True)
    out: list[dict[str, object]] = []
    for sn in wb.sheetnames:
        if not _is_onboard_sheet(sn):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        col_map = _onboard_row_map(rows)
        if not col_map:
            continue
        row_d = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["dwell"])
        row_s = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["signal"])
        row_v = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["speed"])
        row_t = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["travel"])
        row_c = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["crowding"])
        row_tr = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["traffic"])

        for col_name, sess in _ONBOARD_COL_TO_SESSION.items():
            if col_name not in col_map:
                continue
            ci = col_map[col_name]
            bus = _onboard_bus_name_from_notes(rows, ci)
            if study_operator_only and bus and not is_study_bus_company(bus):
                continue
            rec: dict[str, object] = {
                "sheet": sn,
                "session": sess,
                "bus_company": bus,
            }
            if row_d and ci < len(row_d):
                rec["dwell_s"] = parse_dwell_seconds(row_d[ci])
            if row_s and ci < len(row_s):
                rec["signal_s"] = parse_signal_delay_seconds(row_s[ci])
            if row_v and ci < len(row_v):
                rec["speed_mps"] = parse_speed_range_mps(row_v[ci])
            if row_t and ci < len(row_t):
                rec["travel_s"] = parse_travel_time_seconds(row_t[ci])
            if row_c and ci < len(row_c):
                rec["crowding"] = _normalize_crowding_label(str(row_c[ci] or ""))
            if row_tr and ci < len(row_tr):
                rec["traffic"] = _normalize_traffic_label(str(row_tr[ci] or ""))
            out.append(rec)
    wb.close()
    return out


def study_operator_onboard_by_session() -> dict[str, dict[str, object]]:
    """Best on-board record per session for the permissioned operator (San Agustin)."""
    by_sess: dict[str, list[dict[str, object]]] = {}
    for rec in load_onboard_session_records(study_operator_only=True):
        sess = str(rec["session"])
        by_sess.setdefault(sess, []).append(rec)
    chosen: dict[str, dict[str, object]] = {}
    for sess, recs in by_sess.items():
        recs.sort(key=lambda r: str(r.get("sheet", "")))
        chosen[sess] = recs[0]
    return chosen


def erjohn_onboard_by_session() -> dict[str, dict[str, object]]:
    """Deprecated alias — use study_operator_onboard_by_session()."""
    return study_operator_onboard_by_session()


def load_onboard_session_stats(sheet_name: str) -> dict[str, float]:
    """On-board metrics for one session (study operator rows when June workbook is active)."""
    if not FIELD_DATA_PATH.is_file():
        return {}
    if field_workbook_is_june2026():
        rec = study_operator_onboard_by_session().get(sheet_name)
        if rec:
            out: dict[str, float] = {}
            if rec.get("dwell_s"):
                out["onboard_stop_dwell_s"] = float(rec["dwell_s"])
            if rec.get("signal_s"):
                out["onboard_signal_delay_s"] = float(rec["signal_s"])
            if rec.get("speed_mps"):
                out["onboard_cruise_mps"] = float(rec["speed_mps"])
            if rec.get("travel_s"):
                out["onboard_travel_time_s"] = float(rec["travel_s"])
            return out

    wb = openpyxl.load_workbook(FIELD_DATA_PATH, data_only=True)
    dwells: list[float] = []
    signals: list[float] = []
    speeds: list[float] = []
    travels: list[float] = []

    for sn in wb.sheetnames:
        if not _is_onboard_sheet(sn):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        col_map = _onboard_row_map(rows)
        if not col_map:
            continue
        col_idx = None
        for col_name, sess in _ONBOARD_COL_TO_SESSION.items():
            if sess == sheet_name and col_name in col_map:
                col_idx = col_map[col_name]
                break
        if col_idx is None:
            continue

        row_d = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["dwell"])
        row_s = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["signal"])
        row_v = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["speed"])
        row_t = _find_onboard_row(rows, *_ONBOARD_ROW_KEYS["travel"])

        if row_d and col_idx < len(row_d):
            d = parse_dwell_seconds(row_d[col_idx])
            if d and d > 0:
                dwells.append(d)
        if row_s and col_idx < len(row_s):
            d = parse_signal_delay_seconds(row_s[col_idx])
            if d and d > 0:
                signals.append(d)
        if row_v and col_idx < len(row_v):
            v = parse_speed_range_mps(row_v[col_idx])
            if v and v > 0:
                speeds.append(v)
        if row_t and col_idx < len(row_t):
            t = parse_travel_time_seconds(row_t[col_idx])
            if t and t > 0:
                travels.append(t)

    wb.close()
    out = {}
    if dwells:
        out["onboard_stop_dwell_s"] = float(sum(dwells) / len(dwells))
    if signals:
        out["onboard_signal_delay_s"] = float(sum(signals) / len(signals))
    if speeds:
        out["onboard_cruise_mps"] = float(sum(speeds) / len(speeds))
    if travels:
        out["onboard_travel_time_s"] = float(sum(travels) / len(travels))
    return out


def field_bus_profile(
    session: str,
    *,
    profile: str | None = None,
    crowding: str = "",
    traffic: str = "",
) -> SessionBusField | None:
    """Speed band from per-bus traffic tier; session table supplies crowding fallback only."""
    from corridor_config import normalize_traffic_label, traffic_speed_band_kmh

    key = profile or corridor_config.FIELD_ONBOARD_PROFILE
    if key == "jan18":
        row = SESSION_BUS_FIELD.get(session)
    else:
        row = SESSION_BUS_FIELD_SIM.get(session)
    if not row and not traffic:
        return None
    crowd = crowding or (row.crowding if row else "Medium")
    traf = normalize_traffic_label(traffic or (row.traffic if row else "Light"))
    lo, hi = traffic_speed_band_kmh(traf)
    return SessionBusField(lo, hi, crowd, traf)


def apply_onboard_profile(session: str, onboard: dict[str, float], *, profile: str | None = None) -> dict[str, float]:
    """Override Excel aggregates with Jan 18 Sunday on-board row or keep workbook means."""
    key = profile or corridor_config.FIELD_ONBOARD_PROFILE
    if key == "june2026":
        rec = study_operator_onboard_by_session().get(session)
        if rec:
            out = dict(onboard)
            if rec.get("travel_s"):
                out["onboard_travel_time_s"] = float(rec["travel_s"])
            if rec.get("dwell_s"):
                out["onboard_stop_dwell_s"] = float(rec["dwell_s"])
            if rec.get("signal_s"):
                out["onboard_signal_delay_s"] = float(rec["signal_s"])
            if rec.get("speed_mps"):
                out["onboard_cruise_mps"] = float(rec["speed_mps"])
            return out
    if key != "jan18":
        return onboard
    row = SESSION_ONBOARD_JAN18.get(session)
    if not row:
        return onboard
    out = dict(onboard)
    out["onboard_travel_time_s"] = float(row["travel_s"])
    out["onboard_stop_dwell_s"] = float(row["dwell_s"])
    out["onboard_signal_delay_s"] = float(row["signal_s"])
    if "signal_extra_s" in row:
        out["onboard_signal_extra_s"] = float(row["signal_extra_s"])
    return out


def onboard_travel_target_s(
    cal: SessionCalibration,
    *,
    include_informal_stops: bool,
    traffic: str | None = None,
    informal_keys: frozenset[str] | None = None,
) -> float | None:
    """End-to-end on-board target; scales down when fewer informal curbs are encountered."""
    from corridor_config import ALL_INFORMAL_STOP_KEYS, INFORMAL_STOP_KEYS

    base = traffic_onboard_travel_target_s(cal, include_informal_stops=True, traffic=traffic)
    if base is None:
        return None
    if informal_keys is not None:
        n_all = len(INFORMAL_STOP_KEYS)
        frac = len(informal_keys) / n_all if n_all else 0.0
        mult = 0.72 + 0.28 * frac
        return base * mult
    if include_informal_stops:
        return base
    return base * 0.72


def traffic_onboard_signal_target_s(
    cal: SessionCalibration,
    *,
    traffic: str | None = None,
) -> float | None:
    """On-board mean signal wait × traffic tier (Heavy afternoons wait longer at reds)."""
    from corridor_config import TRAFFIC_ONBOARD_SIGNAL_MULT

    raw = cal.onboard_signal_delay_s
    if not raw:
        return None
    from corridor_config import normalize_traffic_label

    tier = normalize_traffic_label(str(traffic or cal.field_traffic or "Light"))
    mult = TRAFFIC_ONBOARD_SIGNAL_MULT.get(tier, 1.0)
    return float(raw) * mult


def traffic_onboard_travel_target_s(
    cal: SessionCalibration,
    *,
    include_informal_stops: bool = True,
    traffic: str | None = None,
) -> float | None:
    """
    On-board 'travel time to destination' target for this bus/session/traffic.

    Uses session on-board mean × traffic tier multiplier.
    """
    from corridor_config import SESSION_ONBOARD_TRAVEL_S, TRAFFIC_ONBOARD_TRAVEL_MULT

    raw = cal.onboard_travel_time_s
    if not raw and cal.sheet in SESSION_ONBOARD_TRAVEL_S:
        raw = SESSION_ONBOARD_TRAVEL_S[cal.sheet]
    if not raw:
        return None
    from corridor_config import normalize_traffic_label

    tier = normalize_traffic_label(str(traffic or cal.field_traffic or "Light"))
    mult = TRAFFIC_ONBOARD_TRAVEL_MULT.get(tier, 1.0)
    target = float(raw) * mult
    if not include_informal_stops:
        target *= 0.72
    return target


def apply_field_speed_to_onboard(session: str, onboard: dict[str, float]) -> dict[str, float]:
    """Override Excel-parsed speed with thesis field bands (24–45 km/h)."""
    june_rec = study_operator_onboard_by_session().get(session) if field_workbook_is_june2026() else None
    crowd = str(june_rec.get("crowding") or "") if june_rec else ""
    traf = str(june_rec.get("traffic") or "") if june_rec else ""
    prof = field_bus_profile(session, crowding=crowd, traffic=traf)
    if not prof:
        return onboard
    out = dict(onboard)
    if field_workbook_is_june2026() and out.get("onboard_cruise_mps"):
        out["field_speed_min_kmh"] = prof.speed_min_kmh
        out["field_speed_max_kmh"] = prof.speed_max_kmh
        return out
    out["onboard_cruise_mps"] = prof.speed_mid_mps
    out["field_speed_min_kmh"] = prof.speed_min_kmh
    out["field_speed_max_kmh"] = prof.speed_max_kmh
    return out


def _parse_number(val) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _header_map(header: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(header):
        key = re.sub(r"\s+", " ", str(h or "").strip().lower())
        if key:
            out[key] = i
    return out


def _col(hmap: dict[str, int], *candidates: str) -> int | None:
    for cand in candidates:
        c = cand.strip().lower()
        if c in hmap:
            return hmap[c]
    for cand in candidates:
        c = cand.strip().lower()
        for k, idx in hmap.items():
            if not k:
                continue
            if c in k or k in c:
                return idx
    return None


def _mean(xs: list[float]) -> float | None:
    return sum(xs) / len(xs) if xs else None


def _mode_str(counter: Counter) -> str:
    if not counter:
        return ""
    top = counter.most_common(5)
    return "; ".join(f"{k}:{v}" for k, v in top if k is not None and str(k).strip() != "")


def _normalize_traffic_label(raw: str) -> str:
    from corridor_config import normalize_traffic_label

    return normalize_traffic_label(raw)


def _normalize_crowding_label(raw: str) -> str:
    from corridor_config import normalize_crowding_label

    return normalize_crowding_label(raw)


def _representative_stationary_level(counter: Counter, *, kind: str) -> str:
    """
    Rush-hour representative level from Bus_Data_Collection stationary rows.

    If Heavy/High appears in >=12% of observations (e.g. morning rush), use that tier
    even when Light/Low is more frequent — matches mixed Low+Heavy field sheets.
    """
    from corridor_config import CROWDING_SEVERITY, TRAFFIC_SEVERITY

    if not counter:
        return "Moderate" if kind == "traffic" else "Medium"
    total = sum(counter.values())
    if kind == "traffic":
        tiers = (
            (("Heavy", "High"), "Heavy"),
            (("Moderate", "Medium"), "Moderate"),
            (("Light", "Low"), "Light"),
        )
        table = TRAFFIC_SEVERITY
        norm = _normalize_traffic_label
    else:
        tiers = (
            (("High", "Heavy"), "High"),
            (("Medium", "Moderate"), "Medium"),
            (("Low", "Light"), "Low"),
        )
        table = CROWDING_SEVERITY
        norm = _normalize_crowding_label

    for keys, label in tiers:
        n = sum(counter.get(k, 0) + counter.get(k.lower(), 0) for k in keys)
        if n / total >= 0.12:
            return label

    table = TRAFFIC_SEVERITY if kind == "traffic" else CROWDING_SEVERITY
    score = sum(table.get(norm(k), 2.0) * v for k, v in counter.items())
    avg = score / total
    if avg >= 2.35:
        return "Heavy" if kind == "traffic" else "High"
    if avg >= 1.65:
        return "Moderate" if kind == "traffic" else "Medium"
    return "Light" if kind == "traffic" else "Low"


@dataclass
class SessionCalibration:
    sheet: str
    data_origin: str
    n_rows_robinsons_location: int
    n_rows_waltermart_location: int
    n_rows_other_location: int
    n_rows_calibration_subset: int
    demand_t0_s: float
    demand_t1_s: float
    demand_span_raw_s: float
    demand_window_capped: bool
    sim_tmax_s: float
    vol_rb_to_wm: int
    vol_wm_to_rb: int
    mean_dwell_rb_s: float | None
    mean_dwell_wm_s: float | None
    mean_boarding_at_origin_s: float | None
    mean_alighting_at_origin_s: float | None
    mean_arrival_rate_at_origin_s: float | None
    crowding_summary: str
    traffic_condition_summary: str
    bus_companies_summary: str
    routes_summary: str
    data_source: str = ""
    # On-Board sheet aggregates (Dec field book — end-to-end trip realism).
    onboard_stop_dwell_s: float | None = None
    onboard_signal_delay_s: float | None = None
    onboard_signal_extra_s: float | None = None
    onboard_cruise_mps: float | None = None
    onboard_travel_time_s: float | None = None
    field_speed_min_kmh: float | None = None
    field_speed_max_kmh: float | None = None
    field_crowding: str = ""
    field_traffic: str = ""


def _load_from_field_workbook(sheet_name: str, *, data_origin: DataOrigin) -> SessionCalibration:
    wb = openpyxl.load_workbook(FIELD_DATA_PATH, data_only=True)
    n_rb = n_wm = n_ot = 0
    drb_all: list[float] = []
    dwm_all: list[float] = []

    subset_times: list[float] = []
    subset_board: list[float] = []
    subset_alight: list[float] = []
    subset_arrival: list[float] = []
    crowd_c = Counter()
    traffic_c = Counter()
    bus_c = Counter()
    route_c = Counter()

    want = "robinsons" if data_origin == "robinsons" else "waltermart"
    n_subset = 0

    for sn in wb.sheetnames:
        if not sn.startswith("Stationary"):
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        raw_header = [str(h).strip() if h is not None else "" for h in rows[1]]
        hmap = _header_map(raw_header)
        i_time = _col(hmap, "time")
        i_bus = _col(hmap, "bus (company name)", "bus")
        i_route = _col(hmap, "route")
        i_loc = _col(hmap, "location (start of oberservation)", "location")
        i_board = _col(hmap, "number of passenger boarding", "boarding")
        i_alight = _col(hmap, "number of passenger alighting", "alighting")
        i_arrival = _col(hmap, "number of passenger arrival rate", "arrival rate")
        i_dwell = _col(hmap, "dwell time")
        i_crowd = _col(hmap, "crowding level", "traffic crowding level", "crowding")
        i_traffic = _col(hmap, "traffic condition")

        if i_time is None or i_loc is None or i_dwell is None:
            continue

        for r in rows[2:]:
            if not r:
                continue
            t_raw = r[i_time]
            sess = None
            if isinstance(t_raw, time):
                sess = _session_from_clock(t_raw)
            else:
                sec = _time_to_seconds(t_raw)
                if sec is not None:
                    sess = _session_from_clock(time(sec // 3600, (sec % 3600) // 60, sec % 60))
            if sess != sheet_name:
                continue

            loc = _normalize_location(str(r[i_loc] or ""))
            if loc == "robinsons":
                n_rb += 1
            elif loc == "waltermart":
                n_wm += 1
            else:
                n_ot += 1

            d = parse_dwell_seconds(r[i_dwell])
            if d and d > 0:
                if loc == "robinsons":
                    drb_all.append(d)
                elif loc == "waltermart":
                    dwm_all.append(d)

            if loc != want:
                continue

            sec = _time_to_seconds(t_raw)
            if sec is None and isinstance(t_raw, time):
                sec = t_raw.hour * 3600 + t_raw.minute * 60 + t_raw.second
            if sec is None:
                continue

            n_subset += 1
            subset_times.append(float(sec))
            if i_board is not None:
                v = _parse_number(r[i_board])
                if v is not None:
                    subset_board.append(v)
            if i_alight is not None:
                v = _parse_number(r[i_alight])
                if v is not None:
                    subset_alight.append(v)
            if i_arrival is not None:
                v = _parse_number(r[i_arrival])
                if v is not None:
                    subset_arrival.append(v)
            if i_crowd is not None and r[i_crowd] is not None:
                crowd_c[str(r[i_crowd]).strip()] += 1
            if i_traffic is not None and r[i_traffic] is not None:
                traffic_c[str(r[i_traffic]).strip()] += 1
            if i_bus is not None and r[i_bus] is not None:
                bus_c[str(r[i_bus]).strip()[:60]] += 1
            if i_route is not None and r[i_route] is not None:
                route_c[str(r[i_route]).strip()[:80]] += 1

    wb.close()

    if not subset_times:
        raise ValueError(
            f"No stationary rows for session {sheet_name!r} at {data_origin!r} in {FIELD_DATA_PATH.name}"
        )

    t_min = min(subset_times)
    obs_rel = [t - t_min for t in subset_times]
    span_raw = max(float(max(obs_rel) - min(obs_rel)), MIN_DEMAND_SPAN_S)
    capped = False
    span_use = span_raw
    if span_raw > MAX_DEMAND_WINDOW_CAP_S:
        span_use = MAX_DEMAND_WINDOW_CAP_S
        capped = True

    origin_label = "robinsons_location" if data_origin == "robinsons" else "waltermart_location"
    onboard = apply_onboard_profile(
        sheet_name,
        apply_field_speed_to_onboard(sheet_name, load_onboard_session_stats(sheet_name)),
    )
    field_crowding = _representative_stationary_level(crowd_c, kind="crowding")
    field_traffic = _representative_stationary_level(traffic_c, kind="traffic")
    prof = field_bus_profile(sheet_name, crowding=field_crowding, traffic=field_traffic)

    return SessionCalibration(
        sheet=sheet_name,
        data_origin=origin_label,
        n_rows_robinsons_location=n_rb,
        n_rows_waltermart_location=n_wm,
        n_rows_other_location=n_ot,
        n_rows_calibration_subset=n_subset,
        demand_t0_s=0.0,
        demand_t1_s=span_use,
        demand_span_raw_s=span_raw,
        demand_window_capped=capped,
        sim_tmax_s=span_use + SIM_CLEARANCE_S,
        vol_rb_to_wm=SESSION_BUS_PER_HOUR_BY_DIRECTION[sheet_name]["rb_to_wm"],
        vol_wm_to_rb=SESSION_BUS_PER_HOUR_BY_DIRECTION[sheet_name]["wm_to_rb"],
        mean_dwell_rb_s=_mean(drb_all),
        mean_dwell_wm_s=_mean(dwm_all),
        mean_boarding_at_origin_s=_mean(subset_board),
        mean_alighting_at_origin_s=_mean(subset_alight),
        mean_arrival_rate_at_origin_s=_mean(subset_arrival),
        crowding_summary=_mode_str(crowd_c),
        traffic_condition_summary=_mode_str(traffic_c),
        bus_companies_summary=_mode_str(bus_c),
        routes_summary=_mode_str(route_c),
        data_source=FIELD_DATA_PATH.name,
        onboard_stop_dwell_s=onboard.get("onboard_stop_dwell_s"),
        onboard_signal_delay_s=onboard.get("onboard_signal_delay_s"),
        onboard_signal_extra_s=onboard.get("onboard_signal_extra_s"),
        onboard_cruise_mps=onboard.get("onboard_cruise_mps"),
        onboard_travel_time_s=onboard.get("onboard_travel_time_s"),
        field_speed_min_kmh=prof.speed_min_kmh if prof else onboard.get("field_speed_min_kmh"),
        field_speed_max_kmh=prof.speed_max_kmh if prof else onboard.get("field_speed_max_kmh"),
        field_crowding=field_crowding,
        field_traffic=field_traffic,
    )


@dataclass
class StationaryVehicleObservation:
    """One bus observed in a Stationary Observation sheet row."""

    obs_id: int
    field_sheet: str
    observation_time: str
    observation_seconds: int
    session: str
    bus_company: str
    route: str
    location: str
    data_origin: str
    boarding: float | None
    alighting: float | None
    arrival_rate: float | None
    dwell_s: float | None
    crowding_raw: str
    traffic_raw: str
    crowding: str
    traffic: str


def _format_clock(seconds: int) -> str:
    seconds = int(seconds) % 86400
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _add_seconds_to_clock(seconds: int, delta_s: float) -> str:
    return _format_clock(int(seconds + delta_s))


def load_all_stationary_vehicles(
    *,
    session: str | None = None,
    data_origin: DataOrigin | None = None,
    study_operator_only: bool | None = None,
    erjohn_only: bool | None = None,
) -> list[StationaryVehicleObservation]:
    """Every bus row from Stationary sheets (San Agustin-only by default on June 2026 workbook)."""
    if study_operator_only is None:
        study_operator_only = erjohn_only
    if study_operator_only is None:
        study_operator_only = field_workbook_is_june2026()
    if not FIELD_DATA_PATH.is_file():
        raise FileNotFoundError(FIELD_DATA_PATH)
    wb = openpyxl.load_workbook(FIELD_DATA_PATH, data_only=True)
    out: list[StationaryVehicleObservation] = []
    obs_id = 0

    for sn in wb.sheetnames:
        if not sn.startswith("Stationary"):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        raw_header = [str(h).strip() if h is not None else "" for h in rows[1]]
        hmap = _header_map(raw_header)
        i_time = _col(hmap, "time")
        i_bus = _col(hmap, "bus (company name)", "bus")
        i_route = _col(hmap, "route")
        i_loc = _col(hmap, "location (start of oberservation)", "location")
        i_board = _col(hmap, "number of passenger boarding", "boarding")
        i_alight = _col(hmap, "number of passenger alighting", "alighting")
        i_arrival = _col(hmap, "number of passenger arrival rate", "arrival rate")
        i_dwell = _col(hmap, "dwell time")
        i_crowd = _col(hmap, "crowding level", "traffic crowding level", "crowding")
        i_traffic = _col(hmap, "traffic condition")
        if i_time is None or i_loc is None:
            continue

        for r in rows[2:]:
            if not r:
                continue
            t_raw = r[i_time]
            sess = None
            sec = None
            if isinstance(t_raw, time):
                sess = _session_from_clock(t_raw)
                sec = t_raw.hour * 3600 + t_raw.minute * 60 + t_raw.second
            else:
                sec = _time_to_seconds(t_raw)
                if sec is not None:
                    sess = _session_from_clock(
                        time(sec // 3600, (sec % 3600) // 60, sec % 60)
                    )
            if sess is None or sec is None:
                continue
            if session is not None and sess != session:
                continue

            loc = _normalize_location(str(r[i_loc] or ""))
            if loc not in ("robinsons", "waltermart"):
                continue
            origin: DataOrigin = "robinsons" if loc == "robinsons" else "waltermart"
            if data_origin is not None and origin != data_origin:
                continue

            crowd_raw = str(r[i_crowd]).strip() if i_crowd is not None and r[i_crowd] else ""
            traffic_raw = str(r[i_traffic]).strip() if i_traffic is not None and r[i_traffic] else ""
            company = str(r[i_bus]).strip() if i_bus is not None and r[i_bus] else ""
            if study_operator_only and not is_study_bus_company(company):
                continue
            obs_id += 1
            out.append(
                StationaryVehicleObservation(
                    obs_id=obs_id,
                    field_sheet=sn,
                    observation_time=_format_clock(sec),
                    observation_seconds=sec,
                    session=sess,
                    bus_company=company,
                    route=str(r[i_route]).strip() if i_route is not None and r[i_route] else "",
                    location=loc,
                    data_origin="robinsons_location" if origin == "robinsons" else "waltermart_location",
                    boarding=_parse_number(r[i_board]) if i_board is not None else None,
                    alighting=_parse_number(r[i_alight]) if i_alight is not None else None,
                    arrival_rate=_parse_number(r[i_arrival]) if i_arrival is not None else None,
                    dwell_s=parse_dwell_seconds(r[i_dwell]) if i_dwell is not None else None,
                    crowding_raw=crowd_raw,
                    traffic_raw=traffic_raw,
                    crowding=_normalize_crowding_label(crowd_raw),
                    traffic=_normalize_traffic_label(traffic_raw),
                )
            )
    wb.close()
    return out


def calibration_for_vehicle(
    obs: StationaryVehicleObservation,
    *,
    data_origin: DataOrigin,
) -> SessionCalibration:
    """Session calibration with this vehicle's traffic, crowding, and dwell from the field row."""
    base = load_session_stats(obs.session, data_origin=data_origin)
    prof = field_bus_profile(
        obs.session,
        crowding=obs.crowding,
        traffic=obs.traffic,
    )
    origin_dwell = obs.dwell_s if obs.dwell_s and obs.dwell_s > 0 else None
    if obs.data_origin == "robinsons_location":
        mean_rb = float(origin_dwell or base.mean_dwell_rb_s or DEFAULT_DWELL_S)
        mean_wm = float(base.mean_dwell_wm_s or DEFAULT_DWELL_S)
    else:
        mean_wm = float(origin_dwell or base.mean_dwell_wm_s or DEFAULT_DWELL_S)
        mean_rb = float(base.mean_dwell_rb_s or DEFAULT_DWELL_S)
    return SessionCalibration(
        sheet=base.sheet,
        data_origin=base.data_origin,
        n_rows_robinsons_location=base.n_rows_robinsons_location,
        n_rows_waltermart_location=base.n_rows_waltermart_location,
        n_rows_other_location=base.n_rows_other_location,
        n_rows_calibration_subset=1,
        demand_t0_s=base.demand_t0_s,
        demand_t1_s=base.demand_t1_s,
        demand_span_raw_s=base.demand_span_raw_s,
        demand_window_capped=base.demand_window_capped,
        sim_tmax_s=base.sim_tmax_s,
        vol_rb_to_wm=base.vol_rb_to_wm,
        vol_wm_to_rb=base.vol_wm_to_rb,
        mean_dwell_rb_s=mean_rb,
        mean_dwell_wm_s=mean_wm,
        mean_boarding_at_origin_s=obs.boarding,
        mean_alighting_at_origin_s=obs.alighting,
        mean_arrival_rate_at_origin_s=obs.arrival_rate,
        crowding_summary=obs.crowding,
        traffic_condition_summary=obs.traffic,
        bus_companies_summary=obs.bus_company,
        routes_summary=obs.route,
        data_source=f"{base.data_source} | vehicle obs {obs.obs_id}",
        onboard_stop_dwell_s=base.onboard_stop_dwell_s,
        onboard_signal_delay_s=traffic_onboard_signal_target_s(base, traffic=obs.traffic)
        or base.onboard_signal_delay_s,
        onboard_signal_extra_s=base.onboard_signal_extra_s,
        onboard_cruise_mps=base.onboard_cruise_mps,
        onboard_travel_time_s=traffic_onboard_travel_target_s(
            base,
            include_informal_stops=True,
            traffic=obs.traffic,
        )
        or base.onboard_travel_time_s,
        field_speed_min_kmh=prof.speed_min_kmh if prof else base.field_speed_min_kmh,
        field_speed_max_kmh=prof.speed_max_kmh if prof else base.field_speed_max_kmh,
        field_crowding=obs.crowding,
        field_traffic=obs.traffic,
    )


def export_stationary_vehicles_csv(path: Path) -> int:
    """Write every stationary bus observation to CSV (field book, one row per bus)."""
    import csv

    from corridor_config import observation_coordinates

    vehicles = load_all_stationary_vehicles()
    if not vehicles:
        return 0
    keys = [
        "obs_id",
        "session",
        "arrived_at_stop_time",
        "bus_company",
        "route",
        "location",
        "coordinates",
        "data_origin",
        "traffic_condition",
        "crowding_level",
        "dwell_at_stop_s",
        "boarding",
        "alighting",
        "arrival_rate",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        for v in vehicles:
            w.writerow(
                {
                    "obs_id": v.obs_id,
                    "session": v.session,
                    "arrived_at_stop_time": v.observation_time,
                    "bus_company": v.bus_company,
                    "route": v.route,
                    "location": v.location,
                    "coordinates": observation_coordinates(
                        data_origin=v.data_origin, location=v.location
                    ),
                    "data_origin": v.data_origin,
                    "traffic_condition": v.traffic,
                    "crowding_level": v.crowding,
                    "dwell_at_stop_s": round(v.dwell_s, 1) if v.dwell_s else "",
                    "boarding": v.boarding if v.boarding is not None else "",
                    "alighting": v.alighting if v.alighting is not None else "",
                    "arrival_rate": v.arrival_rate if v.arrival_rate is not None else "",
                }
            )
    return len(vehicles)


def _trip_direction_for_location(loc: str) -> str | None:
    if loc == "robinsons":
        return "rb_to_wm"
    if loc == "waltermart":
        return "wm_to_rb"
    return None


def stationary_bus_hourly_counts(
    *,
    study_operator_only: bool = False,
) -> dict[str, dict[str, dict[str, int]]]:
    """
    Extrapolate stationary observation counts to buses/hour by session and direction.

    Returns session -> direction -> {study, other, total} hourly rates (rounded ints).
    """
    buckets: dict[str, dict[str, dict[str, list[float]]]] = {
        s: {"rb_to_wm": {"study": [], "other": [], "all": []}, "wm_to_rb": {"study": [], "other": [], "all": []}}
        for s in SESSION_SHEETS
    }
    if not FIELD_DATA_PATH.is_file():
        return {}
    wb = openpyxl.load_workbook(FIELD_DATA_PATH, data_only=True)
    for sn in wb.sheetnames:
        if not sn.startswith("Stationary"):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        raw_header = [str(h).strip() if h is not None else "" for h in rows[1]]
        hmap = _header_map(raw_header)
        i_time = _col(hmap, "time")
        i_bus = _col(hmap, "bus (company name)", "bus")
        i_loc = _col(hmap, "location (start of oberservation)", "location")
        if i_time is None or i_loc is None:
            continue
        for r in rows[2:]:
            if not r:
                continue
            t_raw = r[i_time]
            sess = None
            sec = None
            if isinstance(t_raw, time):
                sess = _session_from_clock(t_raw)
                sec = t_raw.hour * 3600 + t_raw.minute * 60 + t_raw.second
            else:
                sec = _time_to_seconds(t_raw)
                if sec is not None:
                    sess = _session_from_clock(time(sec // 3600, (sec % 3600) // 60, sec % 60))
            if sess is None or sec is None:
                continue
            loc = _normalize_location(str(r[i_loc] or ""))
            direction = _trip_direction_for_location(loc)
            if direction is None:
                continue
            company = str(r[i_bus]).strip() if i_bus is not None and r[i_bus] else ""
            key = "study" if is_study_bus_company(company) else "other"
            if study_operator_only and key != "study":
                continue
            buckets[sess][direction]["all"].append(float(sec))
            buckets[sess][direction][key].append(float(sec))
    wb.close()

    out: dict[str, dict[str, dict[str, int]]] = {}
    for sess, dirs in buckets.items():
        out[sess] = {}
        for direction, parts in dirs.items():
            span_h = 1.0
            if parts["all"]:
                span_s = max(parts["all"]) - min(parts["all"])
                span_h = max(span_s / 3600.0, 0.75)
            out[sess][direction] = {
                "study": max(1, round(len(parts["study"]) / span_h)) if parts["study"] else 0,
                "other": max(0, round(len(parts["other"]) / span_h)),
                "total": max(1, round(len(parts["all"]) / span_h)),
            }
    return out


def other_bus_noise_flow_per_s(session: str, trip_direction: str) -> float:
    """Generic non-permissioned bus flow (veh/s) for mixed-corridor background noise."""
    counts = stationary_bus_hourly_counts()
    row = counts.get(session, {}).get(trip_direction, {})
    hourly = int(row.get("other", 0))
    if hourly <= 0:
        return 0.0
    return hourly / 3600.0


def configure_field_workbook_calibration() -> str:
    """
    Load June 2026 workbook anchors into corridor_config.

    Returns short label for logging (june2026-san-agustin).
    """
    global FIELD_DATA_PATH, DATA_PATH
    FIELD_DATA_PATH = require_june2026_workbook()
    DATA_PATH = FIELD_DATA_PATH

    corridor_config.FIELD_ONBOARD_PROFILE = "june2026"
    study_ob = study_operator_onboard_by_session()
    for sess, rec in study_ob.items():
        if rec.get("travel_s"):
            corridor_config.SESSION_ONBOARD_TRAVEL_S[sess] = float(rec["travel_s"])
        crowd = str(rec.get("crowding") or "")
        traf = str(rec.get("traffic") or "")
        if crowd or traf:
            corridor_config.SESSION_BUS_FIELD_SIM[sess] = SessionBusField(
                *corridor_config.TRAFFIC_FIELD_SPEED_BAND_KMH.get(
                    _normalize_traffic_label(traf or "Light"),
                    corridor_config.TRAFFIC_FIELD_SPEED_BAND_KMH["Light"],
                ),
                _normalize_crowding_label(crowd or "Medium"),
                _normalize_traffic_label(traf or "Light"),
            )

    hourly = stationary_bus_hourly_counts()
    for sess, dirs in hourly.items():
        for direction, parts in dirs.items():
            # Background PH mix anchor uses non-permissioned buses only (Erjohn injected per obs).
            corridor_config.SESSION_BUS_PER_HOUR_BY_DIRECTION[sess][direction] = max(  # type: ignore[index]
                int(parts.get("other", 0)),
                1,
            )

    corridor_config.SESSION_BUS_COUNT_PER_HOUR.update(
        {
            s: corridor_config.SESSION_BUS_PER_HOUR_BY_DIRECTION[s]["rb_to_wm"]
            + corridor_config.SESSION_BUS_PER_HOUR_BY_DIRECTION[s]["wm_to_rb"]
            for s in corridor_config.SESSION_BUS_PER_HOUR_BY_DIRECTION
        }
    )
    corridor_config.SESSION_DEMAND.update(corridor_config.build_session_demand_from_ph_composition())
    return "june2026-san-agustin"


# Auto-configure when June workbook is present.
_ACTIVE_FIELD_LABEL = configure_field_workbook_calibration()


def load_session_stats(sheet_name: str, *, data_origin: DataOrigin) -> SessionCalibration:
    if sheet_name not in SESSION_SHEETS:
        raise ValueError(f"Unknown session {sheet_name!r}; use one of {SESSION_SHEETS}")
    return _load_from_field_workbook(sheet_name, data_origin=data_origin)


def print_calibration(cal: SessionCalibration) -> None:
    src = f" | source={cal.data_source}" if cal.data_source else ""
    print(f"\n--- Calibration: {cal.sheet} | data_origin={cal.data_origin}{src} ---")
    print(
        f"  Location counts (session stationary): Robinsons={cal.n_rows_robinsons_location}, "
        f"Waltermart={cal.n_rows_waltermart_location}, other={cal.n_rows_other_location}"
    )
    print(f"  Rows in this calibration subset: {cal.n_rows_calibration_subset}")
    if cal.demand_window_capped:
        print(
            f"  Demand window: {cal.demand_t0_s:.0f}-{cal.demand_t1_s:.0f} s "
            f"(capped from raw span {cal.demand_span_raw_s:.0f} s)"
        )
    else:
        print(f"  Demand window: {cal.demand_t0_s:.0f}-{cal.demand_t1_s:.0f} s (raw span {cal.demand_span_raw_s:.0f} s)")
    print(f"  Simulation horizon tmax: {cal.sim_tmax_s:.0f} s")
    print(f"  Volumes: RB->WM={cal.vol_rb_to_wm}, WM->RB={cal.vol_wm_to_rb}")
    if cal.mean_dwell_rb_s:
        print(f"  Mean dwell (Robinsons-location rows): {cal.mean_dwell_rb_s:.1f} s")
    if cal.mean_dwell_wm_s:
        print(f"  Mean dwell (Waltermart-location rows): {cal.mean_dwell_wm_s:.1f} s")
    if cal.mean_boarding_at_origin_s is not None:
        print(f"  Mean boarding (subset): {cal.mean_boarding_at_origin_s:.2f}")
    if cal.mean_alighting_at_origin_s is not None:
        print(f"  Mean alighting (subset): {cal.mean_alighting_at_origin_s:.2f}")
    if cal.mean_arrival_rate_at_origin_s is not None:
        print(f"  Mean arrival rate (subset): {cal.mean_arrival_rate_at_origin_s:.2f}")
    if cal.crowding_summary:
        print(f"  Crowding (subset top): {cal.crowding_summary}")
    if cal.traffic_condition_summary:
        print(f"  Traffic condition (subset top): {cal.traffic_condition_summary}")
    if cal.bus_companies_summary:
        print(f"  Bus companies (subset top): {cal.bus_companies_summary}")
    if cal.routes_summary:
        print(f"  Routes (subset top): {cal.routes_summary}")
    if cal.field_speed_max_kmh:
        print(
            f"  Field bus speed: {cal.field_speed_min_kmh:.2f}–{cal.field_speed_max_kmh:.2f} km/h "
            f"(cap {cal.field_speed_max_kmh:.2f} km/h) | crowding={cal.field_crowding} | "
            f"traffic={cal.field_traffic}"
        )
    if cal.onboard_travel_time_s:
        dwell_s = f"{cal.onboard_stop_dwell_s:.0f}s" if cal.onboard_stop_dwell_s else "n/a"
        sig_s = f"{cal.onboard_signal_delay_s:.0f}s" if cal.onboard_signal_delay_s else "n/a"
        print(
            f"  On-board trip: dwell@stop={dwell_s}, signal wait={sig_s}, "
            f"destination={cal.onboard_travel_time_s / 60:.1f} min"
        )


if __name__ == "__main__":
    for sh in SESSION_SHEETS:
        for origin in ("robinsons", "waltermart"):
            try:
                print_calibration(load_session_stats(sh, data_origin=origin))
            except ValueError as e:
                print(f"{sh} / {origin}: {e}")
